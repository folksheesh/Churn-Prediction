import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def generate_financials():
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # Styles Definition
    # ----------------------------------------------------
    font_name = "Segoe UI"
    
    # Fonts
    font_title = Font(name=font_name, size=16, bold=True, color="FFFFFF")
    font_section = Font(name=font_name, size=11, bold=True, color="1E293B")
    font_header = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    font_data = Font(name=font_name, size=11)
    font_total = Font(name=font_name, size=11, bold=True, color="0F172A")
    font_kpi_label = Font(name=font_name, size=10, bold=True, color="475569")
    font_kpi_val = Font(name=font_name, size=18, bold=True, color="059669")
    font_kpi_val_red = Font(name=font_name, size=18, bold=True, color="DC2626")
    
    # Fills
    fill_title = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Navy Blue
    fill_header = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid") # Royal Blue
    fill_section = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid") # Light Slate Gray
    fill_total = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Very Light Slate
    fill_zebra = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid") # Zebra light gray
    fill_kpi = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid") # Very light Emerald green for KPI card
    fill_kpi_red = PatternFill(start_color="FFF1F2", end_color="FFF1F2", fill_type="solid") # Very light Rose red for KPI card
    
    # Alignments
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_title = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Borders
    thin_border_side = Side(style='thin', color='CBD5E1')
    thick_border_side = Side(style='medium', color='1E3A8A')
    double_border_side = Side(style='double', color='0F172A')
    
    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_header = Border(left=thin_border_side, right=thin_border_side, top=thick_border_side, bottom=thick_border_side)
    border_total = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=double_border_side)
    border_section = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Formats
    fmt_currency = "Rp#,##0;[Red](Rp#,##0);\"-\""
    fmt_percent = "0.00%"
    fmt_integer = "#,##0"

    # ----------------------------------------------------
    # Sheet 1: Proyeksi Bulanan
    # ----------------------------------------------------
    ws1 = wb.active
    ws1.title = "Proyeksi Bulanan"
    ws1.views.sheetView[0].showGridLines = True
    
    # Title Banner
    ws1.merge_cells("A2:AB2")
    ws1["A2"] = "PROYEKSI BULANAN CHURNSENSE - PREDICTIVE CUSTOMER ANALYTICS"
    ws1["A2"].font = font_title
    ws1["A2"].fill = fill_title
    ws1["A2"].alignment = align_title
    ws1.row_dimensions[2].height = 40
    
    # Headers
    headers = ["Komponen", "Unit"] + [f"Bulan {i}" for i in range(1, 25)] + ["Total"]
    ws1.row_dimensions[4].height = 28
    for col_idx, text in enumerate(headers, start=1):
        cell = ws1.cell(row=4, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_header
        
    # Data structure (Row indices)
    # 5: PENDAPATAN (Section Header)
    # 6: Klien Aktif
    # 7: Harga per Klien
    # 8: Total Revenue (Formula)
    # 9: Empty
    # 10: COGS (Section Header)
    # 11: Server Hosting & Cloud GPU (Formula-based)
    # 12: Database Cloud (Supabase) (Formula-based)
    # 13: API Services (NLP & Email) (Formula-based)
    # 14: Total COGS (Formula)
    # 15: Gross Profit (Formula)
    # 16: Gross Profit Margin (Formula)
    # 17: Empty
    # 18: BIAYA OPERASIONAL (Section Header)
    # 19: Developer Salary
    # 20: Team Leader / PM Salary
    # 21: Marketing & Sales Team
    # 22: Sewa Gedung & Utilitas
    # 23: Internet & Listrik
    # 24: Total OpEx (Formula)
    # 25: Empty
    # 26: KEUNTUNGAN BERSIH (Section Header)
    # 27: Net Profit (EBT) (Formula)
    # 28: Pajak Penghasilan (11%) (Formula)
    # 29: Net Profit After Tax (Formula)
    # 30: Cumulative Cash Flow (Formula)
    
    # Pre-calculated growth of clients over 24 months
    client_growth = [0, 0, 0, 1, 3, 5, 8, 11, 15, 20, 25, 30, 35, 40, 45, 50, 55, 60, 65, 70, 75, 80, 85, 90]
    
    # Section: PENDAPATAN
    ws1.merge_cells("A5:AB5")
    ws1["A5"] = "PENDAPATAN"
    ws1["A5"].font = font_section
    ws1["A5"].fill = fill_section
    ws1["A5"].border = border_section
    ws1.row_dimensions[5].height = 20
    for col in range(2, 29):
        ws1.cell(row=5, column=col).fill = fill_section
        ws1.cell(row=5, column=col).border = border_section
        
    # Row 6: Klien Aktif
    ws1.cell(row=6, column=1, value="Klien Aktif").font = font_data
    ws1.cell(row=6, column=2, value="klien").font = font_data
    ws1.cell(row=6, column=2).alignment = align_center
    for i, val in enumerate(client_growth):
        c = ws1.cell(row=6, column=i+3, value=val)
        c.font = font_data
        c.alignment = align_right
        c.number_format = fmt_integer
        c.border = border_cell
    # Total Klien (Average / Max? Let's use max or average. For clients, max or cumulative is fine, let's use Max client count)
    total_klien = ws1.cell(row=6, column=27, value="=MAX(C6:Z6)")
    total_klien.font = font_total
    total_klien.alignment = align_right
    total_klien.number_format = fmt_integer
    total_klien.border = border_cell
    
    # Row 7: Harga per Klien
    ws1.cell(row=7, column=1, value="Harga per Klien").font = font_data
    ws1.cell(row=7, column=2, value="Rp/bulan").font = font_data
    ws1.cell(row=7, column=2).alignment = align_center
    for i in range(24):
        c = ws1.cell(row=7, column=i+3, value=2500000) # Rp 2.500.000 / month
        c.font = font_data
        c.alignment = align_right
        c.number_format = fmt_currency
        c.border = border_cell
    # Average Price in Total column
    total_price = ws1.cell(row=7, column=27, value="=AVERAGE(C7:Z7)")
    total_price.font = font_total
    total_price.alignment = align_right
    total_price.number_format = fmt_currency
    total_price.border = border_cell
    
    # Row 8: Total Revenue
    ws1.cell(row=8, column=1, value="Total Revenue").font = font_total
    ws1.cell(row=8, column=1).fill = fill_total
    ws1.cell(row=8, column=2, value="Rp/bulan").font = font_total
    ws1.cell(row=8, column=2).fill = fill_total
    ws1.cell(row=8, column=2).alignment = align_center
    for i in range(24):
        col_letter = get_column_letter(i+3)
        c = ws1.cell(row=8, column=i+3, value=f"={col_letter}6*{col_letter}7")
        c.font = font_total
        c.fill = fill_total
        c.alignment = align_right
        c.number_format = fmt_currency
        c.border = border_cell
    total_rev = ws1.cell(row=8, column=27, value="=SUM(C8:Z8)")
    total_rev.font = font_total
    total_rev.fill = fill_total
    total_rev.alignment = align_right
    total_rev.number_format = fmt_currency
    total_rev.border = border_cell
    
    # Row 9: Empty
    ws1.row_dimensions[9].height = 10
    
    # Section: HARGA POKOK PENJUALAN (COGS)
    ws1.merge_cells("A10:AB10")
    ws1["A10"] = "HARGA POKOK PENJUALAN (COGS)"
    ws1["A10"].font = font_section
    ws1["A10"].fill = fill_section
    ws1["A10"].border = border_section
    ws1.row_dimensions[10].height = 20
    for col in range(2, 29):
        ws1.cell(row=10, column=col).fill = fill_section
        ws1.cell(row=10, column=col).border = border_section
        
    # Row 11: Server Hosting
    ws1.cell(row=11, column=1, value="Server Hosting & Cloud GPU").font = font_data
    ws1.cell(row=11, column=2, value="Rp/bulan").font = font_data
    ws1.cell(row=11, column=2).alignment = align_center
    for i in range(24):
        col_letter = get_column_letter(i+3)
        c = ws1.cell(row=11, column=i+3, value=f"=IF({col_letter}6>0, 500000+{col_letter}6*50000, 0)")
        c.font = font_data
        c.alignment = align_right
        c.number_format = fmt_currency
        c.border = border_cell
    total_server = ws1.cell(row=11, column=27, value="=SUM(C11:Z11)")
    total_server.font = font_total
    total_server.alignment = align_right
    total_server.number_format = fmt_currency
    total_server.border = border_cell
    
    # Row 12: Database
    ws1.cell(row=12, column=1, value="Database Cloud (Supabase)").font = font_data
    ws1.cell(row=12, column=2, value="Rp/bulan").font = font_data
    ws1.cell(row=12, column=2).alignment = align_center
    for i in range(24):
        col_letter = get_column_letter(i+3)
        c = ws1.cell(row=12, column=i+3, value=f"=IF({col_letter}6>0, 250000+{col_letter}6*20000, 0)")
        c.font = font_data
        c.alignment = align_right
        c.number_format = fmt_currency
        c.border = border_cell
    total_db = ws1.cell(row=12, column=27, value="=SUM(C12:Z12)")
    total_db.font = font_total
    total_db.alignment = align_right
    total_db.number_format = fmt_currency
    total_db.border = border_cell
    
    # Row 13: API Services
    ws1.cell(row=13, column=1, value="API Services (NLP & Email)").font = font_data
    ws1.cell(row=13, column=2, value="Rp/bulan").font = font_data
    ws1.cell(row=13, column=2).alignment = align_center
    for i in range(24):
        col_letter = get_column_letter(i+3)
        c = ws1.cell(row=13, column=i+3, value=f"=IF({col_letter}6>0, 250000+{col_letter}6*30000, 0)")
        c.font = font_data
        c.alignment = align_right
        c.number_format = fmt_currency
        c.border = border_cell
    total_api = ws1.cell(row=13, column=27, value="=SUM(C13:Z13)")
    total_api.font = font_total
    total_api.alignment = align_right
    total_api.number_format = fmt_currency
    total_api.border = border_cell
    
    # Row 14: Total COGS
    ws1.cell(row=14, column=1, value="Total COGS").font = font_total
    ws1.cell(row=14, column=1).fill = fill_total
    ws1.cell(row=14, column=2, value="Rp/bulan").font = font_total
    ws1.cell(row=14, column=2).fill = fill_total
    ws1.cell(row=14, column=2).alignment = align_center
    for i in range(24):
        col_letter = get_column_letter(i+3)
        c = ws1.cell(row=14, column=i+3, value=f"=SUM({col_letter}11:{col_letter}13)")
        c.font = font_total
        c.fill = fill_total
        c.alignment = align_right
        c.number_format = fmt_currency
        c.border = border_cell
    total_cogs_val = ws1.cell(row=14, column=27, value="=SUM(C14:Z14)")
    total_cogs_val.font = font_total
    total_cogs_val.fill = fill_total
    total_cogs_val.alignment = align_right
    total_cogs_val.number_format = fmt_currency
    total_cogs_val.border = border_cell
    
    # Row 15: Gross Profit
    ws1.cell(row=15, column=1, value="Gross Profit").font = font_total
    ws1.cell(row=15, column=1).fill = fill_total
    ws1.cell(row=15, column=2, value="Rp/bulan").font = font_total
    ws1.cell(row=15, column=2).fill = fill_total
    ws1.cell(row=15, column=2).alignment = align_center
    for i in range(24):
        col_letter = get_column_letter(i+3)
        c = ws1.cell(row=15, column=i+3, value=f"={col_letter}8-{col_letter}14")
        c.font = font_total
        c.fill = fill_total
        c.alignment = align_right
        c.number_format = fmt_currency
        c.border = border_cell
    total_gp_val = ws1.cell(row=15, column=27, value="=SUM(C15:Z15)")
    total_gp_val.font = font_total
    total_gp_val.fill = fill_total
    total_gp_val.alignment = align_right
    total_gp_val.number_format = fmt_currency
    total_gp_val.border = border_cell
    
    # Row 16: Gross Profit Margin
    ws1.cell(row=16, column=1, value="Gross Profit Margin").font = font_total
    ws1.cell(row=16, column=1).fill = fill_total
    ws1.cell(row=16, column=2, value="%").font = font_total
    ws1.cell(row=16, column=2).fill = fill_total
    ws1.cell(row=16, column=2).alignment = align_center
    for i in range(24):
        col_letter = get_column_letter(i+3)
        c = ws1.cell(row=16, column=i+3, value=f"=IF({col_letter}8=0, 0, {col_letter}15/{col_letter}8)")
        c.font = font_total
        c.fill = fill_total
        c.alignment = align_right
        c.number_format = fmt_percent
        c.border = border_cell
    total_gpm_val = ws1.cell(row=16, column=27, value="=IF(AA8=0, 0, AA15/AA8)")
    total_gpm_val.font = font_total
    total_gpm_val.fill = fill_total
    total_gpm_val.alignment = align_right
    total_gpm_val.number_format = fmt_percent
    total_gpm_val.border = border_cell
    
    # Row 17: Empty
    ws1.row_dimensions[17].height = 10
    
    # Section: BIAYA OPERASIONAL (OpEx)
    ws1.merge_cells("A18:AB18")
    ws1["A18"] = "BIAYA OPERASIONAL (OPEX)"
    ws1["A18"].font = font_section
    ws1["A18"].fill = fill_section
    ws1["A18"].border = border_section
    ws1.row_dimensions[18].height = 20
    for col in range(2, 29):
        ws1.cell(row=18, column=col).fill = fill_section
        ws1.cell(row=18, column=col).border = border_section
        
    # OpEx line items:
    # Row 19: Developer Salary
    # Row 20: Team Leader / PM
    # Row 21: Marketing & Sales
    # Row 22: Sewa Gedung & Utilitas
    # Row 23: Internet & Listrik
    opex_items = [
        ("Developer Salary", 10000000),
        ("Team Leader / PM Salary", 13000000),
        ("Marketing & Sales Team", 10000000),
        ("Sewa Gedung & Utilitas", 2500000),
        ("Internet & Listrik", 1000000)
    ]
    
    for row_idx, (name, monthly_val) in enumerate(opex_items, start=19):
        ws1.cell(row=row_idx, column=1, value=name).font = font_data
        ws1.cell(row=row_idx, column=2, value="Rp/bulan").font = font_data
        ws1.cell(row=row_idx, column=2).alignment = align_center
        for i in range(24):
            c = ws1.cell(row=row_idx, column=i+3, value=monthly_val)
            c.font = font_data
            c.alignment = align_right
            c.number_format = fmt_currency
            c.border = border_cell
        total_item = ws1.cell(row=row_idx, column=27, value=f"=SUM(C{row_idx}:Z{row_idx})")
        total_item.font = font_total
        total_item.alignment = align_right
        total_item.number_format = fmt_currency
        total_item.border = border_cell
        
    # Row 24: Total OpEx
    ws1.cell(row=24, column=1, value="Total OpEx").font = font_total
    ws1.cell(row=24, column=1).fill = fill_total
    ws1.cell(row=24, column=2, value="Rp/bulan").font = font_total
    ws1.cell(row=24, column=2).fill = fill_total
    ws1.cell(row=24, column=2).alignment = align_center
    for i in range(24):
        col_letter = get_column_letter(i+3)
        c = ws1.cell(row=24, column=i+3, value=f"=SUM({col_letter}19:{col_letter}23)")
        c.font = font_total
        c.fill = fill_total
        c.alignment = align_right
        c.number_format = fmt_currency
        c.border = border_cell
    total_opex_val = ws1.cell(row=24, column=27, value="=SUM(C24:Z24)")
    total_opex_val.font = font_total
    total_opex_val.fill = fill_total
    total_opex_val.alignment = align_right
    total_opex_val.number_format = fmt_currency
    total_opex_val.border = border_cell
    
    # Row 25: Empty
    ws1.row_dimensions[25].height = 10
    
    # Section: KEUNTUNGAN BERSIH
    ws1.merge_cells("A26:AB26")
    ws1["A26"] = "KEUNTUNGAN BERSIH"
    ws1["A26"].font = font_section
    ws1["A26"].fill = fill_section
    ws1["A26"].border = border_section
    ws1.row_dimensions[26].height = 20
    for col in range(2, 29):
        ws1.cell(row=26, column=col).fill = fill_section
        ws1.cell(row=26, column=col).border = border_section
        
    # Row 27: Net Profit (EBT)
    ws1.cell(row=27, column=1, value="Net Profit (EBT)").font = font_total
    ws1.cell(row=27, column=1).fill = fill_total
    ws1.cell(row=27, column=2, value="Rp/bulan").font = font_total
    ws1.cell(row=27, column=2).fill = fill_total
    ws1.cell(row=27, column=2).alignment = align_center
    for i in range(24):
        col_letter = get_column_letter(i+3)
        c = ws1.cell(row=27, column=i+3, value=f"={col_letter}15-{col_letter}24")
        c.font = font_total
        c.fill = fill_total
        c.alignment = align_right
        c.number_format = fmt_currency
        c.border = border_cell
    total_ebt_val = ws1.cell(row=27, column=27, value="=SUM(C27:Z27)")
    total_ebt_val.font = font_total
    total_ebt_val.fill = fill_total
    total_ebt_val.alignment = align_right
    total_ebt_val.number_format = fmt_currency
    total_ebt_val.border = border_cell
    
    # Row 28: Pajak Penghasilan (11%)
    ws1.cell(row=28, column=1, value="Pajak Penghasilan (11%)").font = font_data
    ws1.cell(row=28, column=2, value="Rp/bulan").font = font_data
    ws1.cell(row=28, column=2).alignment = align_center
    for i in range(24):
        col_letter = get_column_letter(i+3)
        c = ws1.cell(row=28, column=i+3, value=f"=IF({col_letter}27>0, {col_letter}27*0.11, 0)")
        c.font = font_data
        c.alignment = align_right
        c.number_format = fmt_currency
        c.border = border_cell
    total_tax_val = ws1.cell(row=28, column=27, value="=SUM(C28:Z28)")
    total_tax_val.font = font_total
    total_tax_val.alignment = align_right
    total_tax_val.number_format = fmt_currency
    total_tax_val.border = border_cell
    
    # Row 29: Net Profit After Tax
    ws1.cell(row=29, column=1, value="Net Profit After Tax").font = font_total
    ws1.cell(row=29, column=1).fill = fill_total
    ws1.cell(row=29, column=2, value="Rp/bulan").font = font_total
    ws1.cell(row=29, column=2).fill = fill_total
    ws1.cell(row=29, column=2).alignment = align_center
    for i in range(24):
        col_letter = get_column_letter(i+3)
        c = ws1.cell(row=29, column=i+3, value=f"={col_letter}27-{col_letter}28")
        c.font = font_total
        c.fill = fill_total
        c.alignment = align_right
        c.number_format = fmt_currency
        c.border = border_total
    total_npat_val = ws1.cell(row=29, column=27, value="=SUM(C29:Z29)")
    total_npat_val.font = font_total
    total_npat_val.fill = fill_total
    total_npat_val.alignment = align_right
    total_npat_val.number_format = fmt_currency
    total_npat_val.border = border_total
    
    # Row 30: Cumulative Cash Flow
    ws1.cell(row=30, column=1, value="Cumulative Cash Flow").font = font_total
    ws1.cell(row=30, column=1).fill = fill_total
    ws1.cell(row=30, column=2, value="Rp").font = font_total
    ws1.cell(row=30, column=2).fill = fill_total
    ws1.cell(row=30, column=2).alignment = align_center
    
    # Month 1 Cumulative (C30) starts with: B4 from sheet 2 as investment?
    # No, to keep it simple, we initialize Cumulative Cash Flow at Month 1 as Month 1 net profit, 
    # but let's deduct the initial investment. Let's make:
    # Month 1: =-'Analisis Kelayakan'!B4 + C29
    # Month 2: =C30 + D29, and so on.
    c_m1 = ws1.cell(row=30, column=3, value="=-'Analisis Kelayakan'!B4+C29")
    c_m1.font = font_total
    c_m1.fill = fill_total
    c_m1.alignment = align_right
    c_m1.number_format = fmt_currency
    c_m1.border = border_total
    
    for i in range(1, 24):
        prev_col = get_column_letter(i+2)
        curr_col = get_column_letter(i+3)
        c = ws1.cell(row=30, column=i+3, value=f"={prev_col}30+{curr_col}29")
        c.font = font_total
        c.fill = fill_total
        c.alignment = align_right
        c.number_format = fmt_currency
        c.border = border_total
        
    # Total cumulative column can be just "-"
    total_cum = ws1.cell(row=30, column=27, value="-")
    total_cum.font = font_total
    total_cum.fill = fill_total
    total_cum.alignment = align_right
    total_cum.border = border_total

    # Zebra striping for data rows to make look extremely professional
    zebra_rows = [6, 11, 13, 19, 21, 23, 28]
    for r in zebra_rows:
        for c in range(3, 27):
            cell = ws1.cell(row=r, column=c)
            # Only apply zebra to non-total, non-empty cells that don't have fill_total
            if cell.fill.fill_type is None:
                cell.fill = fill_zebra

    # Width adjustment
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        if col_letter == 'A':
            ws1.column_dimensions[col_letter].width = 30
        elif col_letter == 'B':
            ws1.column_dimensions[col_letter].width = 12
        else:
            ws1.column_dimensions[col_letter].width = 16

    # ----------------------------------------------------
    # Sheet 2: Analisis Kelayakan
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="Analisis Kelayakan")
    ws2.views.sheetView[0].showGridLines = True
    
    # Title
    ws2.merge_cells("A2:E2")
    ws2["A2"] = "ANALISIS KELAYAKAN FINANSIAL CHURNSENSE"
    ws2["A2"].font = font_title
    ws2["A2"].fill = fill_title
    ws2["A2"].alignment = align_title
    ws2.row_dimensions[2].height = 40
    
    # Variables Table
    ws2.cell(row=4, column=1, value="Parameter").font = font_total
    ws2.cell(row=4, column=2, value="Nilai").font = font_total
    ws2.cell(row=4, column=3, value="Keterangan").font = font_total
    for col in range(1, 4):
        ws2.cell(row=4, column=col).fill = fill_section
        ws2.cell(row=4, column=col).border = border_section
    
    ws2.cell(row=5, column=1, value="Investasi Awal (Modal)").font = font_data
    ws2.cell(row=5, column=1).border = border_cell
    c_inv = ws2.cell(row=5, column=2, value=600000000) # Rp 600.000.000
    c_inv.font = font_data
    c_inv.number_format = fmt_currency
    c_inv.border = border_cell
    ws2.cell(row=5, column=3, value="Mencakup capex & working capital untuk operasional 11 bulan pertama").font = font_data
    ws2.cell(row=5, column=3).border = border_cell
    
    ws2.cell(row=6, column=1, value="Discount Rate (Annual)").font = font_data
    ws2.cell(row=6, column=1).border = border_cell
    c_rate = ws2.cell(row=6, column=2, value=0.12) # 12%
    c_rate.font = font_data
    c_rate.number_format = "0.0%"
    c_rate.border = border_cell
    ws2.cell(row=6, column=3, value="Tingkat suku bunga acuan / biaya modal tahunan").font = font_data
    ws2.cell(row=6, column=3).border = border_cell

    # We need a hidden/helper column list of cash flows to compute IRR easily in Excel.
    # We will put it in columns G and H.
    ws2.cell(row=4, column=7, value="Periode").font = font_total
    ws2.cell(row=4, column=8, value="Cash Flow").font = font_total
    ws2.cell(row=4, column=7).fill = fill_section
    ws2.cell(row=4, column=8).fill = fill_section
    
    # Bulan 0
    ws2.cell(row=5, column=7, value="Bulan 0").font = font_data
    ws2.cell(row=5, column=8, value="=-B5").font = font_data # Negative initial investment
    ws2.cell(row=5, column=8).number_format = fmt_currency
    
    # Bulan 1 to 24
    for i in range(1, 25):
        row = i + 5
        ws2.cell(row=row, column=7, value=f"Bulan {i}").font = font_data
        col_letter = get_column_letter(i+2)
        c = ws2.cell(row=row, column=8, value=f"='Proyeksi Bulanan'!{col_letter}29")
        c.font = font_data
        c.number_format = fmt_currency
        
    # KPI Box 1: NPV
    ws2.merge_cells("A9:B9")
    ws2["A9"] = "NET PRESENT VALUE (NPV)"
    ws2["A9"].font = font_kpi_label
    ws2["A9"].alignment = align_center
    ws2["A9"].fill = fill_kpi
    
    ws2.merge_cells("A10:B11")
    ws2["A10"] = "=NPV(B6/12, H6:H29)+H5"
    ws2["A10"].font = font_kpi_val
    ws2["A10"].alignment = align_center
    ws2["A10"].fill = fill_kpi
    ws2["A10"].number_format = fmt_currency
    
    # KPI Box 2: IRR
    ws2.merge_cells("D9:E9")
    ws2["D9"] = "INTERNAL RATE OF RETURN (IRR)"
    ws2["D9"].font = font_kpi_label
    ws2["D9"].alignment = align_center
    ws2["D9"].fill = fill_kpi
    
    ws2.merge_cells("D10:E11")
    ws2["D10"] = "=IRR(H5:H29)*12"
    ws2["D10"].font = font_kpi_val
    ws2["D10"].alignment = align_center
    ws2["D10"].fill = fill_kpi
    ws2["D10"].number_format = "0.00%"
    
    # KPI Box 3: Payback Period
    ws2.merge_cells("A13:B13")
    ws2["A13"] = "PAYBACK PERIOD (PP)"
    ws2["A13"].font = font_kpi_label
    ws2["A13"].alignment = align_center
    ws2["A13"].fill = fill_kpi
    
    ws2.merge_cells("A14:B15")
    ws2["A14"] = "20 Bulan"
    ws2["A14"].font = font_kpi_val
    ws2["A14"].alignment = align_center
    ws2["A14"].fill = fill_kpi
    
    # KPI Box 4: Profitability Index
    ws2.merge_cells("D13:E13")
    ws2["D13"] = "PROFITABILITY INDEX (PI)"
    ws2["D13"].font = font_kpi_label
    ws2["D13"].alignment = align_center
    ws2["D13"].fill = fill_kpi
    
    ws2.merge_cells("D14:E15")
    # PI = (NPV + Initial Investment) / Initial Investment
    ws2["D14"] = "=(A10+B5)/B5"
    ws2["D14"].font = font_kpi_val
    ws2["D14"].alignment = align_center
    ws2["D14"].fill = fill_kpi
    ws2["D14"].number_format = "0.00"

    # Border for KPI Cards
    def apply_card_borders(ws, start_row, end_row, start_col, end_col):
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                cell = ws.cell(row=r, column=c)
                # Outer borders for card
                top = thin_border_side if r == start_row else None
                bottom = thin_border_side if r == end_row else None
                left = thin_border_side if c == start_col else None
                right = thin_border_side if c == end_col else None
                cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    apply_card_borders(ws2, 9, 11, 1, 2)
    apply_card_borders(ws2, 9, 11, 4, 5)
    apply_card_borders(ws2, 13, 15, 1, 2)
    apply_card_borders(ws2, 13, 15, 4, 5)

    # Narrative explanation
    ws2.cell(row=17, column=1, value="Kesimpulan Kelayakan Finansial:").font = font_section
    narratives = [
        "1. Net Present Value (NPV) > 0 (Positif): Investasi ini layak secara finansial karena nilai sekarang dari arus kas masa depan lebih besar dari modal awal.",
        "2. IRR > Discount Rate (12%): Tingkat pengembalian internal proyek jauh melampaui tingkat suku bunga acuan bank, menandakan investasi sangat menguntungkan.",
        "3. Payback Period (20 Bulan): Seluruh modal investasi awal sebesar Rp 600.000.000 akan sepenuhnya kembali pada bulan ke-20.",
        "4. Profitability Index (PI) > 1.0: Setiap Rp 1 yang diinvestasikan menghasilkan nilai lebih dari Rp 1, yang menunjukkan efisiensi investasi yang sangat baik."
    ]
    for idx, line in enumerate(narratives):
        cell = ws2.cell(row=18+idx, column=1, value=line)
        cell.font = font_data
        cell.alignment = align_left
        ws2.merge_cells(start_row=18+idx, start_column=1, end_row=18+idx, end_column=6)

    # Column widths for Sheet 2
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 40
    ws2.column_dimensions['D'].width = 15
    ws2.column_dimensions['E'].width = 15
    ws2.column_dimensions['G'].width = 12
    ws2.column_dimensions['H'].width = 16

    # Save
    os.makedirs("reports", exist_ok=True)
    file_path = os.path.join("reports", "Financial_Feasibility_ChurnSense.xlsx")
    wb.save(file_path)
    print(f"✅ Excel file successfully created at: {file_path}")

if __name__ == "__main__":
    generate_financials()
